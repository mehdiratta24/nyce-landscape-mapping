"""Export the current organizations.json to an Excel spreadsheet matching the
May-15 source layout: one 'Yes' column per capability, comma-joined domains,
'On Website?' derived from engagement_status. 'Coordinates with' is omitted
(removed per Andel's review)."""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JSON_PATH = os.path.join(ROOT, "data", "organizations.json")
OUT_DOWNLOADS = os.path.expanduser("~/Downloads/Data Stewardship landscape analysis - current.xlsx")
OUT_REPO = os.path.join(ROOT, "data", "raw", "landscape-current.xlsx")

CAPABILITIES = [
    "Data Platform",
    "Data Usability & Access",
    "Prioritizing Data",
    "Innovation",
    "Stakeholders: Public & Non Profit",
    "Stakeholders: Research",
    "Stakeholders: Private Sector",
    "Alternative, Proxy Datasets",
    "Domain & Data Expertise",
    "Data Quality & Governance",
    "Coordination",
    "Advocacy & Lobbying",
    "Legal Protection & Litigation",
    "Data Collection & Observing Systems",
    "Data Tools, Products & Models",
    "Integration with Other Data Sources",
]

ORG_TYPE = {
    "nonprofit": "Non Profit",
    "academic": "Academia",
    "company": "Private",
    "government": "Government",
    "independent": "Independent",
}
ON_WEBSITE = {"active": "Yes", "in_contact": "New", "deprioritized": "No"}

with open(JSON_PATH) as f:
    orgs = json.load(f)
orgs.sort(key=lambda o: o["name"].lower())

wb = Workbook()
ws = wb.active
ws.title = "Current"

headers = (
    ["name", "url", "description", "Organization Type"]
    + CAPABILITIES
    + ["Data Domains", "contact_name", "contact_email", "On Website?"]
)
ws.append(headers)

# Header styling
header_fill = PatternFill("solid", start_color="15506C")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = header_fill
    cell.alignment = Alignment(vertical="center", wrap_text=True)

for o in orgs:
    caps = set(o.get("capabilities", []))
    row = [
        o["name"],
        o.get("url", ""),
        o.get("description", ""),
        ORG_TYPE.get(o["organization_type"], o["organization_type"]),
    ]
    row += ["Yes" if c in caps else "" for c in CAPABILITIES]
    row += [
        ", ".join(o.get("dataset_domains", [])),
        o.get("contact_name") or "",
        o.get("contact_email") or "",
        ON_WEBSITE.get(o["engagement_status"], ""),
    ]
    ws.append(row)

# Column widths
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 60
ws.column_dimensions["D"].width = 14
ws.freeze_panes = "A2"

for path in (OUT_DOWNLOADS, OUT_REPO):
    wb.save(path)
    print(f"Wrote {path}")
print(f"  {len(orgs)} organizations, {len(CAPABILITIES)} capability columns")
